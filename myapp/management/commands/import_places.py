from django.core.management.base import BaseCommand
import openpyxl
from myapp.models import Platform, Page, Category, Place
from django.db import transaction



class Command(BaseCommand):
    help = "Loads place data from the provided Excel file"

    def add_arguments(self, parser):
        parser.add_argument('filename', type=str, help='buddy.xlsx')

    def handle(self, *args, **options):
        filename = options['filename']
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        print("Headers:", headers)

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))

            # Platform
            platform_name = data.get('Platform') or 'Unknown'
            platform, _ = Platform.objects.get_or_create(name=platform_name)

            # Page
            page_name = data.get('Page Name') or 'Unknown'
            post_no = str(data.get('Post No', '')).strip() or None
            page, _ = Page.objects.get_or_create(
                name=page_name,
                platform=platform,
                defaults={'post_no': post_no}
            )

            # Category
            category_name = data.get('Category') or 'Unknown'
            category, _ = Category.objects.get_or_create(name=category_name)

            # Place: check for unique (name + page) to avoid duplicates
            name = data.get('Name') or ''
            location = data.get('Location') or ''
            link = data.get('Link') or ''
            place, created = Place.objects.get_or_create(
                name=name,
                page=page,
                category=category,
                defaults={
                    'location': location,
                    'link': link,
                    'views': int(data.get('Views') or 0),
                    'likes': int(data.get('Likes') or 0),
                    'comments_count': int(data.get('Comments') or 0),
                    'google_rating': float(data.get('google rating') or 0) if data.get('google rating') else None,
                    'google_reviews': int(data.get('google reviews') or 0),
                    'sub_type': data.get('Sub Type') or '',
                    'family_friendly': str(data.get('family friendly')).strip().lower() == 'yes',
                    'convenience': data.get('Convinience') or '',
                    'price_per_person': str(data.get('Price Per Person') or ''),
                    'newly_opened': str(data.get('Newly opened flag (< 3months)')).strip().lower() == 'yes',
                    'local_rating': data.get('Local Rating') or '',
                    'pet_friendly': str(data.get('Pet Friendly')).strip().lower() == 'yes',
                    'best_time_to_visit': data.get('Best time to visit') or '',
                    'company_ranking': str(data.get('Company Ranking') or ''),
                    'highlights': data.get('Highlights of place') or '',
                    'top_picks': data.get('Top picks of place') or '',
                    'top_rated_comments': data.get('top rated comments') or '',
                    'recent_comments': data.get('recent comments') or '',
                }
            )

            action_msg = 'Created' if created else 'Already present'
            print(f'{action_msg} place: {place.name} ({category_name}) [{platform_name}]')

        self.stdout.write(self.style.SUCCESS('All rows imported.'))



        # for row in ws.iter_rows(min_row=2, values_only=True):
        #     data = dict(zip(headers, row))
        #     print(data)  # This prints each row as a dict

        # self.stdout.write(self.style.SUCCESS('Successfully read and displayed Excel file rows.'))
